import xlsxwriter,os,shutil
import openpyxl,pyttsx3
import smtplib
from datetime import datetime,date
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from Login import username
from PyQt5.QtCore import QThread,pyqtSignal
from PyQt5.QtWidgets import QMessageBox,QWidget

'''class Worker(QThread):
    finished=pyqtSignal()
    progress=pyqtSignal(int)

    def __init__(self,directory,parent=None):
        self.directory=directory
        super().__init__(parent)

    def run(self):
        today=date.today()
        fromaddr = 'kavyatintin@gmail.com'
        subject='JSSATEB ABSENT NOTIFICATION'
        now = datetime.now()
        current_time = now.strftime("%H:%M:%S")
        wb_obj=openpyxl.load_workbook(os.path.join(self.directory,"attendance.xlsx"))
        sheet=wb_obj.active
        last_empty_row=len(list(sheet.rows))
        for i in range(2,last_empty_row+1):
            if sheet.cell(row=i,column=3).value=='A':
                target=sheet.cell(row=i,column=6).value
                msg = MIMEMultipart()
                msg['From'] = fromaddr
                msg['To'] = target
                msg['Subject'] = subject
                msg.attach(MIMEText("This mail is being sent by the management of JSSATEB. This is to inform that your child "+str(sheet.cell(row=i,column=2).value)+" with usn "+str(sheet.cell(row=i,column=1).value)+" is absent for the class on "+str(today)+" held at "+current_time))
                    
                server = smtplib.SMTP('smtp.gmail.com', 587)
                server.ehlo()
                server.starttls()
                server.ehlo()
                server.login(fromaddr, 'qpyjnhpiadfzxzai')
                server.sendmail(fromaddr, target, msg.as_string())
                server.quit()
        QMessageBox.about(QWidget(), "MAIL SENT", "MAIL SENT SUCCESSFULLY")'''


class SpreadSheetModule():
    def __init__(self,people,usn,branch,sem,section):
        self.people =people
        self.usn=usn
        self.found=False
        self.branch=branch
        self.sem=sem
        self.section=section
        self.network_path=r"\\DESKTOP-B51HC2A\Attendance"

    def createSpreadSheet(self,dir):
        workbook = xlsxwriter.Workbook(os.path.join(dir,"attendance.xlsx"),{'in_memory':True})
        if os.path.exists(os.path.join(self.network_path,username,self.branch,self.sem,self.section))==False:
            print("Does not exist")
            os.makedirs(os.path.join(r"\\DESKTOP-B51HC2A\Attendance",username,self.branch,self.sem,self.section))
        worksheet = workbook.add_worksheet()
 
        worksheet.write('A1', 'USN')
        worksheet.write('B1', 'Name')
        worksheet.write('C1', 'Status')
        worksheet.write('D1', 'Total')
        worksheet.write('E1','Mobile Number')
        worksheet.write('F1',"Parent's Email ID")
        worksheet.write('G1',"Parent's Mobile Number")

        s=0
        r=1
        while s<len(self.people):
            worksheet.write(r,0,self.usn[s])
            worksheet.write(r,1,self.people[s])
            worksheet.write(r,3,0)
            worksheet.write(r,5,'nitinvkavya@gmail.com')
            s+=1
            r+=1
        workbook.close()
        shutil.copy(os.path.join(dir,'attendance.xlsx'),os.path.join(r"\\DESKTOP-B51HC2A\Attendance",username,self.branch,self.sem,self.section,'attendance.xlsx'))

    def updateSpreadSheet(self,peoplePresent,dir):
        #print("from spreadsheet {}".format(peoplePresent
        absentees=[]
        wb_obj=openpyxl.load_workbook(os.path.join(dir,"attendance.xlsx"))
        sheet=wb_obj.active
        last_empty_row=len(list(sheet.rows))
        for i in self.usn:
            self.found=False
            for j in range(2,last_empty_row+1):
                if i==sheet.cell(row=j,column=1).value:
                    self.found=True
                    break
            
            if self.found==False:
                not_init_usn_index=self.usn.index(i)
                not_init_name=self.people[not_init_usn_index]
                sheet.cell(row=j+1,column=1,value=i)
                sheet.cell(row=j+1,column=2,value=not_init_name)
                sheet.cell(row=j+1,column=4,value=1)
                sheet.cell(row=j+1,column=6,value='nitinvkavya@gmail.com')
        wb_obj.save(os.path.join(dir,"attendance.xlsx"))
        shutil.copy(os.path.join(dir,'attendance.xlsx'),os.path.join(r"\\DESKTOP-B51HC2A\Attendance",username,self.branch,self.sem,self.section,'attendance.xlsx'))
        last_empty_row=len(list(sheet.rows))
        if len(peoplePresent)==0:
             for i in range(2,last_empty_row+1):
                sheet.cell(row=i,column=3).value='A'
             engine = pyttsx3.init()
             engine.say("Please confirm the absentees list for today")
             engine.say("The whole class is absent")
             engine.runAndWait()
             wb_obj.save(os.path.join(dir,"attendance.xlsx"))
             shutil.copy(os.path.join(dir,'attendance.xlsx'),os.path.join(r"\\DESKTOP-B51HC2A\Attendance",username,self.branch,self.sem,self.section,'attendance.xlsx'))
            
        else:
            for i in range(2,last_empty_row+1):
                sheet.cell(row=i,column=3).value=''
        #print(last_empty_row)
            for j in peoplePresent:
            #if (sum(peoplePresent[j])/len(peoplePresent[j]))<=55:
                for i in range(2,last_empty_row+1):
                    if sheet.cell(row=i,column=1).value==j.split('-')[0]:
                        sheet.cell(row=i,column=3).value='P'
                        sheet.cell(row=i,column=4).value=sheet.cell(row=i,column=4).value+1
                        break
                for i in range(2,last_empty_row+1):
                    if sheet.cell(row=i,column=3).value=='':
                        sheet.cell(row=i,column=3).value='A'
                        target=sheet.cell(row=i,column=5).value
                        absentees.append(sheet.cell(row=i,column=2).value)
                
            wb_obj.save(os.path.join(dir,"attendance.xlsx"))
            shutil.copy(os.path.join(dir,'attendance.xlsx'),os.path.join(r"\\DESKTOP-B51HC2A\\Attendance",username,self.branch,self.sem,self.section,"attendance.xlsx"))
            engine = pyttsx3.init()
            engine.say("Please confirm the absentees list for today")
            for name in absentees:
                engine.say(name)
                engine.runAndWait()

    def sendMail(self,dir):
        '''self.worker_thread = QThread()
        self.worker = Worker(dir)
        self.worker.moveToThread(self.worker_thread)
        self.worker_thread.started.connect(self.worker.run)
        self.worker.finished.connect(self.worker_thread.quit)       
        self.worker_thread.start()'''
        today=date.today()
        fromaddr = 'kavyatintin@gmail.com'
        subject='JSSATEB ABSENT NOTIFICATION'
        now = datetime.now()
        current_time = now.strftime("%H:%M:%S")
        wb_obj=openpyxl.load_workbook(os.path.join(dir,"attendance.xlsx"))
        sheet=wb_obj.active
        last_empty_row=len(list(sheet.rows))
        for i in range(2,last_empty_row+1):
            if sheet.cell(row=i,column=3).value=='A':
                target=sheet.cell(row=i,column=6).value
                msg = MIMEMultipart()
                msg['From'] = fromaddr
                msg['To'] = target
                msg['Subject'] = subject
                msg.attach(MIMEText("This mail is being sent by the management of JSSATEB. This is to inform that your child "+str(sheet.cell(row=i,column=2).value)+" with usn "+str(sheet.cell(row=i,column=1).value)+" is absent for the class on "+str(today)+" held at "+current_time))
                    
                server = smtplib.SMTP('smtp.gmail.com', 587)
                server.ehlo()
                server.starttls()
                server.ehlo()
                server.login(fromaddr, 'qpyjnhpiadfzxzai')
                server.sendmail(fromaddr, target, msg.as_string())
                server.quit()
        QMessageBox.about(QWidget(), "MAIL SENT", "MAIL SENT SUCCESSFULLY")
        
        


        
        


            
        

        
